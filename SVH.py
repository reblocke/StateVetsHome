# Support Functions for analyzing the state vets home outbreak

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import plotly.graph_objects as go
import scipy.stats as stats


def display_dist(df, label):
    # TODO: separate label from title argument to allow customization
    """takes a df and a column label and graphs the distribution (continuous) for display"""

    # TODO: ensure the height of the box plot is correct

    sns.set(style="white", palette="pastel")
    fig, axes = plt.subplots(2, 1, figsize=(6, 6))
    axes[1].set_aspect(aspect=1)

    # TODO: switch this to displot
    sns.distplot(df[label], ax=axes[0], kde=False, norm_hist=False, color='teal')
    sns.boxplot(data=df, x=label, ax=axes[1], color='skyblue')

    sns.despine(ax=axes[0], top=True, bottom=True, right=True)
    sns.despine(ax=axes[1], top=True, left=True, right=True)

    axes[0].set_xlabel("")
    axes[0].set_ylabel("Count per bin", fontsize='large')

    row_label = "{lab}\nMean: {mean:.1f}, Std Dev: {std:.1f}\nMedian: {med:.1f}, IQR: [{lower:.1f}, {upper:.1f}]\nCount: {count:.0f}"\
        .format(lab=label, mean=df[label].describe()['mean'], std=df[label].describe()['std'],
                med=df[label].describe()['50%'], lower=df[label].describe()['25%'], upper=df[label].describe()['75%'],
                count=df[label].describe()['count'])

    axes[1].set_xlabel(row_label, fontsize='large')
    axes[1].get_shared_x_axes().join(axes[1], axes[0])
    axes[1].set(xlim=(0, None))

    fig.suptitle("Distribution of: " + str(label), fontsize='xx-large')
    fig.tight_layout(rect=[0, 0, 1, .9])  # .95 to leave space for title
    fig.savefig('dist figs/Display Dist ' + str(label) + '.png', dpi=100)
    plt.close()


def display_cats(df, label):
    # TODO: separate label from title argument to allow customization
    """takes a df and a column label and graphs the counts of categories for display"""
    sns.set(style="white", palette="cubehelix")
    fig, axes = plt.subplots(2, 1, figsize=(6, 6))

    sns.histplot(df[label].dropna(), ax=axes[0])
    counts = df[label].value_counts()
    axes[1].pie(counts, labels=counts.index)
    axes[1].text(-3.0, -1.3, "Value counts:\n\n"+str(counts), fontsize=12)

    sns.despine(ax=axes[0], top=True, bottom=True, right=True, left=True)
    axes[0].set_xlabel("")

    fig.suptitle("Distribution of: " + str(label), fontsize='xx-large')
    # fig.tight_layout(rect=[0, 0, 1, .9])  # .95 to leave space for title
    fig.savefig('dist figs/Display Cat ' + str(label) + '.png', dpi=100)
    plt.close()


def file_to_df(location):
    """loads the database excel into a pandas dataframe"""
    db = pd.read_excel(location, sheet_name=0, header=0, index_col=0)

    # category definitions. Others are just string equivalents, unordered
    palliative_cat = pd.api.types.CategoricalDtype(categories=['N', 'Y'], ordered=True)
    p_acp_cat = pd.api.types.CategoricalDtype(categories=['GOC', 'AD', 'POLST', 'LST'], ordered=True)
    O2_cat = pd.api.types.CategoricalDtype(categories=['None', 'NC', 'HFNC'], ordered=True)
    setting_cat = pd.api.types.CategoricalDtype(categories=['Acute care', 'ICU'], ordered=True)
    discharge_o2_cat = pd.api.types.CategoricalDtype(categories=['N/A', 'No', 'Yes'], ordered=True)
    acp_surr_cat = pd.api.types.CategoricalDtype(categories=['Unknown', 'POLST', 'Surrogate', 'Patient'], ordered=True)
    code_cat = pd.api.types.CategoricalDtype(categories=['Unknown', 'Full code', 'DNR', 'DNR/DNI'], ordered=True)

    # convert data types
    db['Age'] = db['Age'].astype('int')
    db['Gender'] = db['Gender'].astype('category')
    db['BMI'] = db['BMI'].astype('float')
    db['Ethnicity'] = db['Ethnicity'].astype('category')
    db['Death'] = db['Death'].astype('category')
    db['Oxygen Delivery'] = db['Oxygen Delivery'].astype(O2_cat)
    db['Setting'] = db['Setting'].astype(setting_cat)
    db['New Discharge O2'] = db['New Discharge O2'].replace({"N": "No", "Y": "Yes", None: "N/A"})\
        .astype(discharge_o2_cat)
    db['LOS'] = db['LOS'].astype('int')
    # db['Admit']
    # db['Discharge']

    # -- Advanced care planning
    db['Palliative Consult'] = db['Palliative Consult'].astype(palliative_cat)
    db['CCI'] = db['CCI'].astype('int')
    db['Prior ACP type'] = db['Prior ACP type'].astype(p_acp_cat)
    db['Prior Decision Maker'] = db['Prior Decision Maker'].replace({"N": "No", "Y": "Yes", None: "Yes"})\
        .astype(acp_surr_cat)
    # db['Date of prior ACP']
    db['Prior Code status'] = db['Prior Code status'].replace(
        {None: "Unknown"}).astype(code_cat)
    db['Prior limitations on artificial nutrition'] = db['Prior limitations on artificial nutrition'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    db['Prior limitations on intubation'] = db['Prior limitations on intubation'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    db['Prior limitations of ICU transfer'] = db['Prior limitations of ICU transfer'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    db['Ok for IV fluids or antibiotics'] = db['Ok for IV fluids or antibiotics'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    db['Ok for long term nutrition or intubation'] = db['Ok for long term nutrition or intubation'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    db['Prior Comfort care'] = db['Prior Comfort care'].replace(
        {"X": "Yes", None: "No"}).astype('category')
    # Prior Stated goals
    db['Hospitalization ACP'] = db['Hospitalization ACP'].astype('category')
    db['Current Decision Maker'] = db['Current Decision Maker'].astype(acp_surr_cat)
    db['Change from prior decision maker'] = db['Change from prior decision maker'].astype('category')
    # Date of LST
    db['Code Status At Hospitalization'] = db['Code Status At Hospitalization'].astype(code_cat)
    db['Comfort care'] = db['Comfort care'].replace({"X": "Yes", None: "No"}).astype('category')
    db['ICU transfer acceptable to patient?'] = db['ICU transfer acceptable to patient?'].replace(
        {"N": "No", "Y": "Yes", None: "Yes"}).astype('category')
    db['Change in code status from prior ACP on admission'] = db['Change in code status from prior ACP on admission']\
        .replace({"N": "No", "Y": "Yes", None: "N/A"}).astype('category')
    db['Direct of Change in code status on admit'] = db['Direct of Change in code status on admit']\
        .replace({"Less": "Less", "More": "More", None: "No Change or N/A"}).astype('category')
    # TODO: integrate two columns to be Yes-more, yes-less, no change, not known
    db['Subsequent changes during hospitalization'] = db['Subsequent changes during hospitalization'].replace(
        {"N": "No", "Y": "Yes", None: "No"}).astype('category')
    # Decision Maker Subsequent
    # Change to more or less agreesive measures Subsequent
    # TODO: figure how to do a 'last' code status

    # TODO: do columns for symptom onset, test date, admit date

    # ---Symptoms----
    db['Symptoms prior to admit'] = db['Symptoms prior to admit'].astype('category')
    db['Fever'] = db['Fever'].replace({'X':'Yes', None:'No'}).astype('category')
    db['SOB'] = db['SOB'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Cough'] = db['Cough'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Sinus Congestion'] = db['Sinus Congestion'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Malaise or Fatigue or Weakness'] = db['Malaise or Fatigue or Weakness'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Diarrhea'] = db['Diarrhea'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Confusion'] = db['Confusion'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Anorexia'] = db['Anorexia'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Myalgias Athralgias'] = db['Myalgias Athralgias'].replace({'X':'Yes', None:'No'}).astype('category')
    db['HA'] = db['HA'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Sore throat'] = db['Sore throat'].replace({'X':'Yes', None:'No'}).astype('category')
    db['Abdominal pain'] = db['Abdominal pain'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['SIRS criteria met on admission'] = db['SIRS criteria met on admission'].replace({'Y': 'Yes', "N": 'No'}).astype('category')

    # ---Admission Signs----
    db['Temp'] = db['Temp'].astype('float')
    db['SBP'] = db['SBP'].astype('float')
    db['DBP'] = db['DBP'].astype('float')
    db['Pulse'] = db['Pulse'].astype('float')
    db['RR'] = db['RR'].astype('float')
    db['O2'] = db['O2'].astype('float')
    db['WBC'] = db['WBC'].astype('float')
    db['SIRS criteria met'] = db['SIRS criteria met'].astype('int')
    db['Supp O2'] = db['Supp O2'].astype('int')

    #db['Days after admit test positive'] = db['Date of + test'] - db['Admit']
    db['Days after admit test positive'] = db.apply(calc_time_diff_days, axis=1, args=('Date of + test', 'Admit'))

    # ---Comorbidities----
    db['MI'] = db['MI'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['CHF'] = db['CHF'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['PVD'] = db['PVD'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['CVA or TIA'] = db['CVA or TIA'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Dementia'] = db['Dementia'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['COPD'] = db['COPD'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Connective tissue disease'] = db['Connective tissue disease'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['PUD'] = db['PUD'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Liver disease'] = db['Liver disease'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['DM'] = db['DM'].replace({'Complicated': 'Yes/Complicated', 'Uncomplicated': 'Yes/Uncomplicated', None: 'No'}).astype('category')
    db['Mod-Sev CKD'] = db['Mod-Sev CKD'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Solid tumor'] = db['Solid tumor'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Leukemia'] = db['Leukemia'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Lymphoma'] = db['Lymphoma'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['AIDS'] = db['AIDS'].replace({'X': 'Yes', None: 'No'}).astype('category')
    db['Hemiplegia'] = db['Hemiplegia'].replace({'X': 'Yes', None: 'No'}).astype('category')

    return db


def make_tables(db):
    """creates tables and exports to excel"""

    #Subgroups: current decision maker (patient vs surrogate/polst) - stratified by patient vs other decision-maker
    # TODO: Dead v not? ICU vs acute care, current decision maker (patient vs surrogate/polst)
    patient_dec_db = db.loc[db['Current Decision Maker'] == 'Patient']
    surr_dec_db = db.loc[db['Current Decision Maker'] != 'Patient']

    num_total = len(db.index)
    num_patient_dec = len(patient_dec_db)
    num_surr_dec = len(surr_dec_db)

    column_labels = ['All Patients\n(n='+str(num_total)+')',
                     'Decision-maker: Patient\n(n='+str(num_patient_dec)+')',
                     'Decision-maker: Surrogate\n(n='+str(num_surr_dec)+')']
    workbook = Workbook()

    # Table 1: Demographics and Pre-Hosp
    demo_row_labels = []
    demographics = []

    # Age
    demo_row_labels.append('AGE')
    demographics.append(
        (std_string(db['Age'].describe()),
         std_string(patient_dec_db['Age'].describe()),
         std_string(surr_dec_db['Age'].describe()))
    )
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # Gender
    demo_row_labels.append('GENDER')
    demographics.append(("", "", ""))

    for gender in db['Gender'].value_counts().keys():
        demo_row_labels.append(gender)
        demographics.append((count_string_indiv(db['Gender'].value_counts()[gender], num_total),
                             count_string_indiv(patient_dec_db['Gender'].value_counts(dropna=False)[gender], num_patient_dec),
                             count_string_indiv(surr_dec_db['Gender'].value_counts(dropna=False)[gender], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # Ethnicity
    demo_row_labels.append('ETHNICITY')
    demographics.append(("", "", ""))

    for gender in db['Ethnicity'].value_counts().keys():
        demo_row_labels.append(gender)
        demographics.append((count_string_indiv(db['Ethnicity'].value_counts()[gender], num_total),
                             count_string_indiv(patient_dec_db['Ethnicity'].value_counts(dropna=False)[gender], num_patient_dec),
                             count_string_indiv(surr_dec_db['Ethnicity'].value_counts(dropna=False)[gender], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # BMI
    demo_row_labels.append('BMI')
    demographics.append(
        (iqr_string(db['BMI'].describe()),
         iqr_string(patient_dec_db['BMI'].describe()),
         iqr_string(surr_dec_db['BMI'].describe()))
    )
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # COMORBIDITIES
    demo_row_labels.append('COMORBIDITIES')
    demographics.append(("", "", ""))
    # CCI
    demo_row_labels.append('TOTAL CHARLSON\nCOMORBIDITY INDEX')
    demographics.append(
        (iqr_string(db['CCI'].describe()),
         iqr_string(patient_dec_db['CCI'].describe()),
         iqr_string(surr_dec_db['CCI'].describe()))
    )

    demo_row_labels.append('CAD')
    demographics.append(("", "", ""))
    for cad in db['MI'].value_counts().keys():
        demo_row_labels.append(cad)
        demographics.append((count_string_indiv(db['MI'].value_counts()[cad], num_total),
                             count_string_indiv(patient_dec_db['MI'].value_counts(dropna=False)[cad], num_patient_dec),
                             count_string_indiv(surr_dec_db['MI'].value_counts(dropna=False)[cad], num_surr_dec)))
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    demo_row_labels.append('CHF')
    demographics.append(("", "", ""))
    for chf in db['CHF'].value_counts().keys():
        demo_row_labels.append(chf)
        demographics.append((count_string_indiv(db['CHF'].value_counts()[chf], num_total),
                             count_string_indiv(patient_dec_db['CHF'].value_counts(dropna=False)[chf], num_patient_dec),
                             count_string_indiv(surr_dec_db['CHF'].value_counts(dropna=False)[chf], num_surr_dec)))
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    #'PVD'

    demo_row_labels.append('CVA or TIA')
    demographics.append(("", "", ""))
    for cva in db['CVA or TIA'].value_counts().keys():
        demo_row_labels.append(cva)
        demographics.append((count_string_indiv(db['CVA or TIA'].value_counts()[cva], num_total),
                             count_string_indiv(patient_dec_db['CVA or TIA'].value_counts(dropna=False)[cva], num_patient_dec),
                             count_string_indiv(surr_dec_db['CVA or TIA'].value_counts(dropna=False)[cva], num_surr_dec)))
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    demo_row_labels.append('Dementia')
    demographics.append(("", "", ""))
    for dem in db['Dementia'].value_counts().keys():
        demo_row_labels.append(dem)
        demographics.append((count_string_indiv(db['Dementia'].value_counts()[dem], num_total),
                             count_string_indiv(patient_dec_db['Dementia'].value_counts(dropna=False)[dem],
                                                num_patient_dec),
                             count_string_indiv(surr_dec_db['Dementia'].value_counts(dropna=False)[dem],
                                                num_surr_dec)))
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    #'COPD'
    #'Connective tissue disease'
    #'PUD'
    #'Liver disease'

    demo_row_labels.append('DM')
    demographics.append(("", "", ""))
    for dm in db['DM'].value_counts().keys():
        demo_row_labels.append(dm)
        demographics.append((count_string_indiv(db['DM'].value_counts()[dm], num_total),
                             count_string_indiv(patient_dec_db['DM'].value_counts(dropna=False)[dm],
                                                num_patient_dec),
                             count_string_indiv(surr_dec_db['DM'].value_counts(dropna=False)[dm],
                                                num_surr_dec)))
    demo_row_labels.append("")
    demographics.append(("", "", ""))

    #'Mod-Sev CKD'
    #'Solid tumor'
    #'Leukemia'
    #'Lymphoma'
    #'AIDS'
    #'Hemiplegia'

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # Prior ACP type
    demo_row_labels.append('PRIOR ADVANCED\nCARE PLANNING')
    demographics.append(("", "", ""))

    for acp in db['Prior ACP type'].value_counts().keys():
        demo_row_labels.append(acp)
        demographics.append((count_string_indiv(db['Prior ACP type'].value_counts()[acp], num_total),
                             count_string_indiv(patient_dec_db['Prior ACP type'].value_counts(dropna=False)[acp], num_patient_dec),
                             count_string_indiv(surr_dec_db['Prior ACP type'].value_counts(dropna=False)[acp], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # 'Prior Decision Maker'
    demo_row_labels.append('PRIOR DECISION-\nMAKER')
    demographics.append(("", "", ""))

    for decider in db['Prior Decision Maker'].value_counts().keys():
        demo_row_labels.append(decider)
        demographics.append((count_string_indiv(db['Prior Decision Maker'].value_counts()[decider], num_total),
                             count_string_indiv(patient_dec_db['Prior Decision Maker'].value_counts(dropna=False)[decider], num_patient_dec),
                             count_string_indiv(surr_dec_db['Prior Decision Maker'].value_counts(dropna=False)[decider], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # 'Prior Code status'
    demo_row_labels.append('PRIOR\nCODE STATUS')
    demographics.append(("", "", ""))

    for status in db['Prior Code status'].value_counts().keys():
        demo_row_labels.append(status)
        demographics.append((count_string_indiv(db['Prior Code status'].value_counts()[status], num_total),
                             count_string_indiv(patient_dec_db['Prior Code status'].value_counts(dropna=False)[status], num_patient_dec),
                             count_string_indiv(surr_dec_db['Prior Code status'].value_counts(dropna=False)[status], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    # 'Prior Comfort care'
    demo_row_labels.append('PRIOR\nCOMFORT CARE')
    demographics.append(("", "", ""))

    for cc in db['Prior Comfort care'].value_counts().keys():
        demo_row_labels.append(cc)
        demographics.append((count_string_indiv(db['Prior Comfort care'].value_counts()[cc], num_total),
                             count_string_indiv(patient_dec_db['Prior Comfort care'].value_counts(dropna=False)[cc], num_patient_dec),
                             count_string_indiv(surr_dec_db['Prior Comfort care'].value_counts(dropna=False)[cc], num_surr_dec)))

    demo_row_labels.append("")
    demographics.append(("", "", ""))

    demographics_df = pd.DataFrame(demographics, columns=column_labels, index=demo_row_labels)
    demographic_worksheet = workbook.worksheets[0]
    demographic_worksheet.title = "Demographics and Pre-hosp"

    for r in dataframe_to_rows(demographics_df, index=True, header=True):
        demographic_worksheet.append(r)

    for cell in demographic_worksheet['A'] + demographic_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

    # -------------------------
    # Table 2 - Presentation
    column_labels = ['All Patients\n(n=' + str(num_total) + ')',
                     'Decision-maker: Patient\n(n=' + str(num_patient_dec) + ')',
                     'Decision-maker: Surrogate\n(n=' + str(num_surr_dec) + ')']
    pres_row_labels = []
    presentation = []

    pres_row_labels.append('DAYS AFTER ADMIT COVID+ TEST')
    presentation.append((iqr_string(db['Days after admit test positive'].describe()),
         iqr_string(patient_dec_db['Days after admit test positive'].describe()),
         iqr_string(surr_dec_db['Days after admit test positive'].describe())))

    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('SYMPTOMS PRIOR TO ADMIT?')
    presentation.append(("", "", ""))
    for x in db['Symptoms prior to admit'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Symptoms prior to admit'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Symptoms prior to admit'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Symptoms prior to admit'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    # ---Symptoms----
    pres_row_labels.append('SYMPTOMS AT PRESENTATION')
    presentation.append(("", "", ""))

    pres_row_labels.append('Fever')
    presentation.append(("", "", ""))
    for x in db['Fever'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Fever'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Fever'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Fever'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Shortness of Breath')
    presentation.append(("", "", ""))
    for x in db['SOB'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['SOB'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['SOB'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['SOB'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Cough')
    presentation.append(("", "", ""))
    for x in db['Cough'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Cough'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Cough'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Cough'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Sinus Congestion')
    presentation.append(("", "", ""))
    for x in db['Sinus Congestion'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Sinus Congestion'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Sinus Congestion'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Sinus Congestion'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Malaise or Fatigue or Weakness')
    presentation.append(("", "", ""))
    for x in db['Malaise or Fatigue or Weakness'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Malaise or Fatigue or Weakness'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Malaise or Fatigue or Weakness'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Malaise or Fatigue or Weakness'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Diarrhea')
    presentation.append(("", "", ""))
    for x in db['Diarrhea'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Diarrhea'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Diarrhea'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Diarrhea'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Confusion')
    presentation.append(("", "", ""))
    for x in db['Confusion'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Confusion'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Confusion'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Confusion'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Anorexia')
    presentation.append(("", "", ""))
    for x in db['Anorexia'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Anorexia'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Anorexia'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Anorexia'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Myalgias Athralgias')
    presentation.append(("", "", ""))
    for x in db['Myalgias Athralgias'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Myalgias Athralgias'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Myalgias Athralgias'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Myalgias Athralgias'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('HA')
    presentation.append(("", "", ""))
    for x in db['HA'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['HA'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['HA'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['HA'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Sore throat')
    presentation.append(("", "", ""))
    for x in db['Sore throat'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Sore throat'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Sore throat'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Sore throat'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Abdominal pain')
    presentation.append(("", "", ""))
    for x in db['Abdominal pain'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['Abdominal pain'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Abdominal pain'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Abdominal pain'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    # ---Admission Signs----
    pres_row_labels.append('SIGNS AT PRESENTATION')
    presentation.append(("", "", ""))

    pres_row_labels.append('2+ SIRS criteria met on admission')
    presentation.append(("", "", ""))
    for x in db['SIRS criteria met on admission'].value_counts().keys():
        pres_row_labels.append(x)
        presentation.append((count_string_indiv(db['SIRS criteria met on admission'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['SIRS criteria met on admission'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['SIRS criteria met on admission'].value_counts(dropna=False)[x], num_surr_dec)))
    pres_row_labels.append("")
    presentation.append(("", "", ""))

    pres_row_labels.append('Number of SIRS criteria met')
    presentation.append((iqr_string(db['SIRS criteria met'].describe()),
         iqr_string(patient_dec_db['SIRS criteria met'].describe()),
         iqr_string(surr_dec_db['SIRS criteria met'].describe())))

    pres_row_labels.append('Temperature')
    presentation.append((std_string(db['Temp'].describe()),
         std_string(patient_dec_db['Temp'].describe()),
         std_string(surr_dec_db['Temp'].describe())))

    pres_row_labels.append('SBP')
    presentation.append((std_string(db['SBP'].describe()),
         std_string(patient_dec_db['SBP'].describe()),
         std_string(surr_dec_db['SBP'].describe())))

    pres_row_labels.append('DBP')
    presentation.append((std_string(db['DBP'].describe()),
         std_string(patient_dec_db['DBP'].describe()),
         std_string(surr_dec_db['DBP'].describe())))

    pres_row_labels.append('Pulse')
    presentation.append((std_string(db['Pulse'].describe()),
         std_string(patient_dec_db['Pulse'].describe()),
         std_string(surr_dec_db['Pulse'].describe())))

    pres_row_labels.append('Respiratory Rate')
    presentation.append((std_string(db['RR'].describe()),
         std_string(patient_dec_db['RR'].describe()),
         std_string(surr_dec_db['RR'].describe())))

    pres_row_labels.append('SpO2')
    presentation.append((std_string(db['O2'].describe()),
         std_string(patient_dec_db['O2'].describe()),
         std_string(surr_dec_db['O2'].describe())))

    pres_row_labels.append('Supplemental O2 (L/min)')
    presentation.append((iqr_string(db['Supp O2'].describe()),
         iqr_string(patient_dec_db['Supp O2'].describe()),
         iqr_string(surr_dec_db['Supp O2'].describe())))

    pres_row_labels.append('WBC')
    presentation.append((std_string(db['WBC'].describe()),
         std_string(patient_dec_db['WBC'].describe()),
         std_string(surr_dec_db['WBC'].describe())))

    presentation_df = pd.DataFrame(presentation, columns=column_labels, index=pres_row_labels)
    presentation_worksheet = workbook.create_sheet(title="Presentation", index=1)

    for r in dataframe_to_rows(presentation_df, index=True, header=True):
        presentation_worksheet.append(r)

    for cell in presentation_worksheet['A'] + presentation_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    # -------------------------
    # Table 3 - Hospitalization  / Outcome

    column_labels = ['All Patients\n(n=' + str(num_total) + ')',
                     'Decision-maker: Patient\n(n=' + str(num_patient_dec) + ')',
                     'Decision-maker: Surrogate\n(n=' + str(num_surr_dec) + ')']
    hosp_row_labels = []
    hospitalization = []

    # 'Setting'
    hosp_row_labels.append('MAXIMAL CARE SETTING')
    hospitalization.append(("", "", ""))
    for x in db['Setting'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Setting'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Setting'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Setting'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Oxygen Delivery'
    hosp_row_labels.append('MAXIMAL OXYGEN DELIVERY MODE')
    hospitalization.append(("", "", ""))
    for x in db['Oxygen Delivery'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Oxygen Delivery'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Oxygen Delivery'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Oxygen Delivery'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Code Status At Hospitalization'
    hosp_row_labels.append('CODE STATUS\nAT HOSPITALIZATION')
    hospitalization.append(("", "", ""))
    for x in db['Code Status At Hospitalization'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Code Status At Hospitalization'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Code Status At Hospitalization'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Code Status At Hospitalization'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Current Decision Maker'
    hosp_row_labels.append('DECISION-MAKER\nAT HOSPITALIZATION')
    hospitalization.append(("", "", ""))
    for x in db['Current Decision Maker'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Current Decision Maker'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Current Decision Maker'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Current Decision Maker'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Hospitalization ACP'
    hosp_row_labels.append('ADVANCED CARE PLANNING TYPE\nAT HOSPITALIZATION')
    hospitalization.append(("", "", ""))
    for x in db['Hospitalization ACP'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Hospitalization ACP'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Hospitalization ACP'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Hospitalization ACP'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Change from prior decision maker'
    hosp_row_labels.append('CHANGE FROM PRIOR\nSURROGATE DECISION-MAKER?')
    hospitalization.append(("", "", ""))
    for x in db['Change from prior decision maker'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Change from prior decision maker'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Change from prior decision maker'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Change from prior decision maker'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Change in code status from prior ACP on admission'
    # hosp_row_labels.append('CHANGE IN CODE STATUS\nFROM PRIOR?')
    # hospitalization.append(("", "", ""))
    # for x in db['Change in code status from prior ACP on admission'].value_counts().keys():
    #    hosp_row_labels.append(x)
    #    hospitalization.append((count_string_indiv(db['Change in code status from prior ACP on admission'].value_counts()[x], num_total),
    #                         count_string_indiv(patient_dec_db['Change in code status from prior ACP on admission'].value_counts(dropna=False)[x], num_patient_dec),
    #                         count_string_indiv(surr_dec_db['Change in code status from prior ACP on admission'].value_counts(dropna=False)[x], num_surr_dec)))

    # hosp_row_labels.append("")
    # hospitalization.append(("", "", ""))

    # 'Direct of Change in code status on admit'
    hosp_row_labels.append('DIRECTION OF CHANGE\nIN ADMIT CODE STATUS\nCOMPARED TO PRIOR')
    hospitalization.append(("", "", ""))
    for x in db['Direct of Change in code status on admit'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Direct of Change in code status on admit'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Direct of Change in code status on admit'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Direct of Change in code status on admit'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Subsequent changes during hospitalization'
    hosp_row_labels.append('SUBSEQUENT CHANGES IN GOALS\n DURING HOSPITALIZATION')
    hospitalization.append(("", "", ""))
    for x in db['Subsequent changes during hospitalization'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Subsequent changes during hospitalization'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Subsequent changes during hospitalization'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Subsequent changes during hospitalization'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Comfort care'
    hosp_row_labels.append('COMFORT CARE \nAT HOSPITALIZATION')
    hospitalization.append(("", "", ""))
    for x in db['Comfort care'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Comfort care'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Comfort care'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Comfort care'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'ICU transfer acceptable to patient?'
    hosp_row_labels.append('ICU TRANSFER ACCEPTABLE?')
    hospitalization.append(("", "", ""))
    for x in db['ICU transfer acceptable to patient?'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['ICU transfer acceptable to patient?'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['ICU transfer acceptable to patient?'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['ICU transfer acceptable to patient?'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Palliative Consult'
    hosp_row_labels.append('PALLIATIVE CONSULT?')
    hospitalization.append(("", "", ""))
    for x in db['Palliative Consult'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Palliative Consult'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Palliative Consult'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Palliative Consult'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'LOS'
    hosp_row_labels.append('LENGTH OF STAY')
    hospitalization.append(
        (iqr_string(db['LOS'].describe()),
         iqr_string(patient_dec_db['LOS'].describe()),
         iqr_string(surr_dec_db['LOS'].describe()))
    )
    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'New Discharge O2'
    hosp_row_labels.append('NEW DISCHARGE\nOXYGEN REQUIREMENT')
    hospitalization.append(("", "", ""))
    for x in db['New Discharge O2'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['New Discharge O2'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['New Discharge O2'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['New Discharge O2'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    # 'Death'
    hosp_row_labels.append('DEATH')
    hospitalization.append(("", "", ""))
    for x in db['Death'].value_counts().keys():
        hosp_row_labels.append(x)
        hospitalization.append((count_string_indiv(db['Death'].value_counts()[x], num_total),
                             count_string_indiv(patient_dec_db['Death'].value_counts(dropna=False)[x], num_patient_dec),
                             count_string_indiv(surr_dec_db['Death'].value_counts(dropna=False)[x], num_surr_dec)))

    hosp_row_labels.append("")
    hospitalization.append(("", "", ""))

    hospitalization_df = pd.DataFrame(hospitalization, columns=column_labels, index=hosp_row_labels)
    hospitalization_worksheet = workbook.create_sheet(title="Hospitalization and Course", index=2)

    for r in dataframe_to_rows(hospitalization_df, index=True, header=True):
        hospitalization_worksheet.append(r)

    for cell in hospitalization_worksheet['A'] + hospitalization_worksheet[1]:
        cell.style = 'Pandas'
        cell.alignment = Alignment(wrapText=True, vertical='center')

    # TODO: add final code status / ACP info

    workbook.save("tables.xlsx")
    return


def iqr_string(summary):
    """returns string of 'mean [IQR 25,75], n=_' when given a dataframe.describe() result
    for non-normal dist data"""
    output = "".join(['%.1f' % summary['mean'], " [IQR ", '%.1f, ' % summary['25%'], '%.1f]' % summary['75%']])
    return output


def std_string(summary):
    """returns string of 'mean +/- std, n=_' when given a dataframe.describe() result
    for normal dist data"""
    output = "".join(['%.1f' % summary['mean'], " (+/- ", '%.1f)' % summary['std']])
    return output


def count_string(counts_series, num_patients):
    """returns string of the counts of each from a dataframe.value_counts() result and num_patients, which is the total
    number of patients (not observations, e.g. in the case of patients with multiple comorbidities, so that percentages
    of the patients can be calculated"""
    output = ""
    for label in counts_series.keys():
        output += label + " = %.0f" % counts_series[label]
        percent = (counts_series[label] / num_patients) * 100
        output += ' (%.1f%%)\n' % percent
    return output[:-1]  # take off the final \n


def count_string_indiv(num, num_patients):
    """returns an string with the number and percentage of an individuals value"""
    output = "%.0f/" % num
    output += str(num_patients)
    if num_patients is not 0:
        percentage = (num / num_patients) * 100
    else:
        percentage = 0.0
    output += ' (%.1f%%)' % percentage
    return output


# TODO: Add yes/no string


def calc_time_diff_days(row, date1, date2):
    """take a row and 2 dates and calculate the difference in time between date 1 and date 2 expressed in days
    must take columns (date1) and (date2) that contain datetime objects"""
    diff = row[date1] - row[date2]
    return diff.days


def statistical_tests(df):
    """limited statistical tests comparing groups of: needing surrogate vs not"""

    patient_dec_df = df.loc[df['Current Decision Maker'] == 'Patient']
    surr_dec_df = df.loc[df['Current Decision Maker'] != 'Patient']

    num_total = len(df.index)
    num_patient_dec = len(patient_dec_df)
    num_surr_dec = len(surr_dec_df)

    print("\nAge")
    print(stats.ttest_ind(patient_dec_df['Age'], surr_dec_df['Age']))
    print("\nGender")
    print([patient_dec_df['Gender'].value_counts(), surr_dec_df['Gender'].value_counts()])
    print(stats.fisher_exact([patient_dec_df['Gender'].value_counts(), surr_dec_df['Gender'].value_counts()]))
    #Note: ethnicity required Fisher-Freemon-Halton r x c generalization of exact test
    print("\nBMI")
    print(stats.ttest_ind(patient_dec_df['BMI'], surr_dec_df['BMI']))
    print("\nCCI")
    print(stats.ttest_ind(patient_dec_df['CCI'], surr_dec_df['CCI']))
    print("\nCAD")
    print([patient_dec_df['MI'].value_counts(sort=False), surr_dec_df['MI'].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df['MI'].value_counts(sort=False), surr_dec_df['MI'].value_counts(sort=False)]))
    print("\nCHF")
    print([patient_dec_df['CHF'].value_counts(sort=False), surr_dec_df['CHF'].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df['CHF'].value_counts(sort=False), surr_dec_df['CHF'].value_counts(sort=False)]))
    print("\nCVA or TIA")
    print([patient_dec_df['CVA or TIA'].value_counts(sort=False), surr_dec_df['CVA or TIA'].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df['CVA or TIA'].value_counts(sort=False), surr_dec_df['CVA or TIA'].value_counts(sort=False)]))
    print("\nDementia")
    print([patient_dec_df['Dementia'].value_counts(sort=False), surr_dec_df['Dementia'].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df['Dementia'].value_counts(sort=False), surr_dec_df['Dementia'].value_counts(sort=False)]))
    #DM done in stata
    #Prior ACP done in stata
    print("\nDays after admit pos")
    print(stats.ttest_ind(patient_dec_df["Days after admit test positive"], surr_dec_df["Days after admit test positive"]))
    print("Symptoms prior to admit")
    print([patient_dec_df["Symptoms prior to admit"].value_counts(sort=False), surr_dec_df["Symptoms prior to admit"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Symptoms prior to admit"].value_counts(sort=False), surr_dec_df["Symptoms prior to admit"].value_counts(sort=False)]))
    print("Fever")
    print([patient_dec_df["Fever"].value_counts(sort=False),
           surr_dec_df["Fever"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Fever"].value_counts(sort=False),
                              surr_dec_df["Fever"].value_counts(sort=False)]))
    print("SOB")
    print([patient_dec_df["SOB"].value_counts(sort=False),
           surr_dec_df["SOB"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["SOB"].value_counts(sort=False),
                              surr_dec_df["SOB"].value_counts(sort=False)]))
    print("Cough")
    print([patient_dec_df["Cough"].value_counts(sort=False),
           surr_dec_df["Cough"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Cough"].value_counts(sort=False),
                              surr_dec_df["Cough"].value_counts(sort=False)]))
    print("Sinus Congestion")
    print([patient_dec_df["Sinus Congestion"].value_counts(sort=False),
           surr_dec_df["Sinus Congestion"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Sinus Congestion"].value_counts(sort=False),
                              surr_dec_df["Sinus Congestion"].value_counts(sort=False)]))
    print("Malaise or Fatigue or Weakness")
    print([patient_dec_df["Malaise or Fatigue or Weakness"].value_counts(sort=False),
           surr_dec_df["Malaise or Fatigue or Weakness"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Malaise or Fatigue or Weakness"].value_counts(sort=False),
                              surr_dec_df["Malaise or Fatigue or Weakness"].value_counts(sort=False)]))
    print("Diarrhea")
    print([patient_dec_df["Diarrhea"].value_counts(sort=False),
           surr_dec_df["Diarrhea"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Diarrhea"].value_counts(sort=False),
                              surr_dec_df["Diarrhea"].value_counts(sort=False)]))
    print("Confusion")
    print([patient_dec_df["Confusion"].value_counts(sort=False),
           surr_dec_df["Confusion"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Confusion"].value_counts(sort=False),
                              surr_dec_df["Confusion"].value_counts(sort=False)]))
    print("Anorexia")
    print([patient_dec_df["Anorexia"].value_counts(sort=False),
           surr_dec_df["Anorexia"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Anorexia"].value_counts(sort=False),
                              surr_dec_df["Anorexia"].value_counts(sort=False)]))
    print("Myalgias Athralgias")
    print([patient_dec_df["Myalgias Athralgias"].value_counts(sort=False),
           surr_dec_df["Myalgias Athralgias"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Myalgias Athralgias"].value_counts(sort=False),
                              surr_dec_df["Myalgias Athralgias"].value_counts(sort=False)]))
    print("HA")
    print([patient_dec_df["HA"].value_counts(sort=False),
           surr_dec_df["HA"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["HA"].value_counts(sort=False),
                              surr_dec_df["HA"].value_counts(sort=False)]))
    #Skipped Sore throat and Abdominal pain
    print("Sore throat")
    print([patient_dec_df["Sore throat"].value_counts(sort=False),
           surr_dec_df["Sore throat"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Sore throat"].value_counts(sort=False),
                              surr_dec_df["Sore throat"].value_counts(sort=False)]))
    print("Abdominal pain")
    print([patient_dec_df["Abdominal pain"].value_counts(sort=False),
           surr_dec_df["Abdominal pain"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["Abdominal pain"].value_counts(sort=False),
                              surr_dec_df["Abdominal pain"].value_counts(sort=False)]))
    print("SIRS criteria met on admission")
    print([patient_dec_df["SIRS criteria met on admission"].value_counts(sort=False),
           surr_dec_df["SIRS criteria met on admission"].value_counts(sort=False)])
    print(stats.fisher_exact([patient_dec_df["SIRS criteria met on admission"].value_counts(sort=False),
                              surr_dec_df["SIRS criteria met on admission"].value_counts(sort=False)]))
    print("\nSIRS criteria met")
    print(stats.ttest_ind(patient_dec_df["SIRS criteria met"], surr_dec_df["SIRS criteria met"]))
    print("\nTemp")
    print(stats.ttest_ind(patient_dec_df["Temp"], surr_dec_df["Temp"]))
    print("\nSBP")
    print(stats.ttest_ind(patient_dec_df["SBP"], surr_dec_df["SBP"]))
    print("\nDBP")
    print(stats.ttest_ind(patient_dec_df["DBP"], surr_dec_df["DBP"]))
    print("\nPulse")
    print(stats.ttest_ind(patient_dec_df["Pulse"], surr_dec_df["Pulse"]))
    print("\nRR")
    print(stats.ttest_ind(patient_dec_df["RR"], surr_dec_df["RR"]))
    print("\nO2")
    print(stats.ttest_ind(patient_dec_df["O2"], surr_dec_df["O2"]))
    print("\nSupp O2")
    print(stats.ttest_ind(patient_dec_df["Supp O2"], surr_dec_df["Supp O2"]))
    print("\nWBC")
    print(stats.ttest_ind(patient_dec_df["WBC"].dropna(), surr_dec_df["WBC"].dropna()))
    print("\nLOS")
    print(stats.ttest_ind(patient_dec_df["LOS"].dropna(), surr_dec_df["LOS"].dropna()))

    return


def code_status_alluvial(df):
    """takes the df and creates a sankey diagram (alluvial - with time points oriented vertically) to show change in
    code status at 3 time points:

    NOTE: at the moment, manually generating the widths by filling in
    TODO: automatically generate these
    """
    fig = go.Figure(data=[go.Sankey(
        node = dict(
          pad= 15,
          thickness= 45,
          line= dict(color = "black", width = 0.5),
          label= ["Prior: Unknown\nn=3", "Prior: Full Code\nn=5", "Prior: DNR only\nn=3", "Prior: DNI/DNR\nn=14",
                   "Admit: Full Code\nn=3", "Admit: DNR only\nn=1", "Admit: DNI/DNR\nn=21",
                   "Discharge: Full Code\nn=4", "Discharge: DNR only\nn=0", "Discharge: DNI/DNR\nn=20", "Death in Hosp\nn=1",
                   "Alive\nn=18", "Dead within 30d\nn=7"],
          color= ['#CBB4D5', '#6C91C4', '#FEF3C7', '#EBBAB5', # #EBBAB5 = red, #FEF3C7 = yellow, #6C91C4 = blue
                   '#6C91C4', '#FEF3C7', '#EBBAB5',
                   '#6C91C4', '#FEF3C7', '#EBBAB5', '#808B96',
                   '#EBE9E9', '#808B96']
        ),
        link = dict(
            source=[0, 0, 0, 1, 1, 1, 2, 2, 2, 3, 3, 3, 4, 4, 4, 5, 5, 5, 6, 6, 6, 6, 7, 7, 8, 8, 9, 9, 10, 10],
            target=[4, 5, 6, 4, 5, 6, 4, 5, 6, 4, 5, 6, 7, 8, 9, 7, 8, 9, 7, 8, 9, 10, 11, 12, 11, 12, 11, 12, 11, 12],
            value=[0, 0, 3, 2, 1, 2, 1, 0, 2, 0, 0, 14, 3, 0, 0, 0, 0, 1, 1, 0, 20, 1, 4, 0, 0, 0, 14, 7, 0, 1]
        ),
        arrangement='perpendicular',
        ids = ['a', 'b', 'c', 'd',
               'e', 'f', 'g',
               'h', 'i', 'j', 'k', 'l']
    )])

    fig.update_layout(font_size=25)
    fig.show()


def visualizations(db):
    """takes the database, creates visualizations of the data"""
    display_dist(db, 'Age')
    display_cats(db, 'Gender')
    display_dist(db, 'BMI')
    display_cats(db, 'Ethnicity')
    display_cats(db, 'Death')
    display_cats(db, 'Oxygen Delivery')
    display_cats(db, 'Setting')
    display_cats(db, 'New Discharge O2')
    display_dist(db, 'LOS')
    display_cats(db, 'Palliative Consult')
    display_dist(db, 'CCI')
    display_cats(db, 'Prior ACP type')
    display_cats(db, 'Prior Decision Maker')
    display_cats(db, 'Prior Code status')
    display_cats(db, 'Prior limitations on artificial nutrition')
    display_cats(db, 'Prior limitations on intubation')
    display_cats(db, 'Prior limitations of ICU transfer')
    display_cats(db, 'Ok for IV fluids or antibiotics')
    display_cats(db, 'Prior limitations of ICU transfer')
    display_cats(db, 'Ok for long term nutrition or intubation')
    display_cats(db, 'Prior Comfort care')
    display_cats(db, 'Hospitalization ACP')
    display_cats(db, 'Current Decision Maker')
    display_cats(db, 'Change from prior decision maker')
    display_cats(db, 'Code Status At Hospitalization')
    display_cats(db, 'Comfort care')
    display_cats(db, 'ICU transfer acceptable to patient?')
    display_cats(db, 'Change in code status from prior ACP on admission')
    display_cats(db, 'Direct of Change in code status on admit')
    display_cats(db, 'Subsequent changes during hospitalization')

    display_cats(db, 'Symptoms prior to admit')
    display_cats(db, 'Fever')
    display_cats(db, 'SOB')
    display_cats(db, 'Sinus Congestion')
    display_cats(db, 'Malaise or Fatigue or Weakness')
    display_cats(db, 'Diarrhea')
    display_cats(db, 'Confusion')
    display_cats(db, 'Anorexia')
    display_cats(db, 'Myalgias Athralgias')
    display_cats(db, 'HA')
    display_cats(db, 'Sore throat')
    display_cats(db, 'Abdominal pain')
    display_cats(db, 'SIRS criteria met on admission')

    display_dist(db, 'Temp')
    display_dist(db, 'SBP')
    display_dist(db, 'DBP')
    display_dist(db, 'Pulse')
    display_dist(db, 'RR')
    display_dist(db, 'O2')
    display_dist(db, 'WBC')
    display_dist(db, 'SIRS criteria met')
    display_dist(db, 'Supp O2')
    display_dist(db, 'Age')
    display_dist(db, 'Days after admit test positive')

    display_cats(db, 'MI')
    display_cats(db, 'CHF')
    display_cats(db, 'PVD')
    display_cats(db, 'CVA or TIA')
    display_cats(db, 'Dementia')
    display_cats(db, 'COPD')
    display_cats(db, 'Connective tissue disease')
    display_cats(db, 'PUD')
    display_cats(db, 'Liver disease')
    display_cats(db, 'DM')
    display_cats(db, 'Mod-Sev CKD')
    display_cats(db, 'Solid tumor')
    display_cats(db, 'Leukemia')
    display_cats(db, 'Lymphoma')
    display_cats(db, 'AIDS')
    display_cats(db, 'Hemiplegia')


def main():
    db_loc = "/Users/reblocke/Box/Residency Personal Files/Scholarly Work/SVH COVID Outbreak/Database/WorkingDb.xls"
    db = file_to_df(db_loc)
    statistical_tests(db)
    db.to_excel('output.xlsx')

    make_tables(db)
    visualizations(db)
    code_status_alluvial(db)

    # Hypotheses? Statistical tests?



    # Difference of direction between aggressiveness of surrogate changes vs not
    # Likelihood of change whether patient or surrogate is making decision at time of admission.
    # Difference in direction based on comorbidity index

    '''
    Other points from roxanne's email
    
    Did prior LST/POLST specify any limitations of care? (outside of DNR, DNI)
    -No artificial nutrition - 2
    -Limited trial of intubation - 2
    -Limited trial of tube feeds - 4
    None of these had any impact on clinical course.
    
    How many had a change in their wishes? (regardless of decision maker)
    - On admission - 6
    - During admission - 3
    - Both on admission & during admission - 2
    - Of those, did this change patient outcome? 3 (2 moved to comfort care due to critical illness, 1 not critically ill but stopped further medical interventions and discharged)
    '''



if __name__ == '__main__':
    main()